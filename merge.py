import argparse
import os
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook

from name import NameMode
from utils import copy_sheet


parser = argparse.ArgumentParser(
    description="Merges single sheets from multiple workbooks into a single workbook"
)
parser.add_argument(
    "-d", "--dir", type=str, required=True, help="Source workbook directory"
)
parser.add_argument(
    "-t", "--target", type=str, required=True, help="Target workbook filepath"
)
parser.add_argument(
    "-n",
    "--name-mode",
    type=NameMode,
    required=False,
    default=NameMode.SEQUENTIAL,
    help="Target sheet name mode",
)
parser.add_argument(
    "-c",
    "--name-cell",
    type=str,
    required=False,
    default=None,
    help='Source sheet name cell for naming target sheet (if using "from_cell" name mode)',
)

args = parser.parse_args()
workbook_dir = Path(args.dir)
target_workbook_filepath = Path(args.target)
name_mode = args.name_mode
name_cell = args.name_cell

if not workbook_dir.is_dir():
    raise Exception("Workbook directory passed is not a directory.")

if not workbook_dir.exists():
    raise Exception("Workbook directory does not exist.")

if target_workbook_filepath.suffix != "":
    if target_workbook_filepath.suffix != ".xlsx":
        raise Exception("Invalid target workbook file extension.")
else:
    target_workbook_filepath = Path(
        target_workbook_filepath.parent, f"{target_workbook_filepath.name}.xlsx"
    )

if name_mode == NameMode.FROM_CELL:
    if name_cell is None:
        raise Exception(
            'Source sheet name cell should be provided when target sheet name mode is set to "from_cell".'
        )
    elif not re.match(r"[A-z]+\d+", name_cell):
        raise Exception("Source sheet name cell is invalid.")

os.makedirs(target_workbook_filepath.parent, exist_ok=True)

workbook_filepaths = [p for p in workbook_dir.iterdir() if p.suffix == ".xlsx"]

if len(workbook_filepaths) == 0:
    raise Exception(f'No .xlsx files found in "{workbook_dir}".')

target_workbook_filepath = target_workbook_filepath
target_wb = Workbook()
source_sheet_names = set()

for i, workbook_filepath in enumerate(workbook_filepaths):
    source_wb = load_workbook(workbook_filepath)
    source_ws = source_wb.active
    if name_mode == NameMode.FROM_CELL:
        target_sheet_name = source_ws[name_cell].value
        pass
    elif name_mode == NameMode.FROM_SHEET_NAME:
        target_sheet_name = source_wb.sheetnames[0]
        if target_sheet_name in source_sheet_names:
            raise Exception(
                f'Source sheet name already exists (duplicate): "{target_sheet_name}".'
            )
        source_sheet_names.add(target_sheet_name)
    else:
        target_sheet_name = f"Sheet {i + 1}"
    target_ws = target_wb.create_sheet(target_sheet_name)
    copy_sheet(source_ws, target_ws)

if "Sheet" in target_wb.sheetnames:
    target_wb.remove(target_wb["Sheet"])

target_wb.save(target_workbook_filepath)
